import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Count
from scanner.models import OMRSubmission

def build_rankings_excel(submissions, file_object):
    """
    Generates a password-protected, read-only Excel workbook with two sheets:
    1. Standings: Student scores and metrics (including Multi-marked column).
    2. Evaluator Summary: Sheets checked by each evaluator.
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Standings
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Standings"
    
    # Header Styles
    header_fill = PatternFill(start_color="0D2B4E", end_color="0D2B4E", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    headers = [
        'S.No.',
        'School Code',
        'School Name',
        'Full Roll Number',
        'Student Name',
        'Category',
        'Question Paper Set',
        'Marks Obtained',
        'Percentage',
        'Correct Answers',
        'Wrong Answers',
        'Unanswered Questions',
        'Multi-marked Questions',
        'Scan Status',
        'Scan Timestamp',
        'Operator',
        'Evaluator'
    ]
    
    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    # Write Data
    for row_idx, s in enumerate(submissions, 2):
        p = s.participant
        r = getattr(s, 'result', None)
        
        school_code = p.school.code if (p and p.school) else '—'
        school_name = p.school.name if (p and p.school) else '—'
        roll_number = p.roll_number if p else '—'
        student_name = p.student_name if (p and p.student_name) else '—'
        category = p.get_group_display() if p else '—'
        paper_set = p.get_paper_set_display() if p else '—'
        evaluator = s.evaluator_name if s.evaluator_name else '—'
        
        if r:
            marks_obtained = r.score
            percentage = f"{r.percentage}%"
            correct = r.score
            unanswered = r.unanswered_count
            multi_marked = r.multi_marked_count
            wrong = 50 - r.score - r.unanswered_count - r.multi_marked_count
        else:
            marks_obtained = '—'
            percentage = '—'
            correct = '—'
            wrong = '—'
            unanswered = '—'
            multi_marked = '—'
            
        status = s.get_status_display()
        timestamp = s.uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
        operator = s.operator.username if s.operator else 'System'
        
        row_values = [
            row_idx - 1,
            school_code,
            school_name,
            roll_number,
            student_name,
            category,
            paper_set,
            marks_obtained,
            percentage,
            correct,
            wrong,
            unanswered,
            multi_marked,
            status,
            timestamp,
            operator,
            evaluator
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            
            # Preserve leading zeros for school code and roll number as String type
            if col_idx in [2, 4] and val != '—':
                cell.value = str(val)
                cell.data_type = 's'
                cell.number_format = '@'
                cell.alignment = center_align
            else:
                cell.value = val
                if isinstance(val, int):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
    # Adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Enable Sheet Protection (Read-Only)
    ws1.protection.sheet = True
    ws1.protection.password = "bkj_qms_2026"
    
    # ----------------------------------------------------
    # Sheet 2: Evaluator Summary
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Evaluator Summary")
    
    eval_headers = ['Evaluator Name', 'Sheets Checked']
    for col_idx, header in enumerate(eval_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    # Query evaluator totals
    evaluators = OMRSubmission.objects.filter(
        status='EVALUATED', 
        is_accepted=True
    ).values('evaluator_name').annotate(count=Count('id')).order_by('-count')
    
    for row_idx, ev in enumerate(evaluators, 2):
        name = ev['evaluator_name'] if ev['evaluator_name'] else 'Unknown/Operator'
        count = ev['count']
        
        c1 = ws2.cell(row=row_idx, column=1, value=name)
        c1.alignment = left_align
        
        c2 = ws2.cell(row=row_idx, column=2, value=count)
        c2.alignment = center_align
        
    # Adjust column widths
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18
    
    # Enable Sheet Protection (Read-Only)
    ws2.protection.sheet = True
    ws2.protection.password = "bkj_qms_2026"
    
    # Save Workbook to file object
    wb.save(file_object)
